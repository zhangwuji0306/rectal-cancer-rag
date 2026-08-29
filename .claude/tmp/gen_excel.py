# -*- coding: utf-8 -*-
"""Generate Excel report: pending-conversion list, full-library classification, missing refs."""
import json, pathlib, sys, csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")
ROOT = pathlib.Path("E:/writing-rag")
data = json.loads((ROOT / ".claude/tmp/mapping.json").read_text(encoding="utf-8"))
files = data["files"]
converted_stems = {p.stem for p in (ROOT / "converted").glob("*.md")}

# live conversion status
for f in files:
    f["conv_now"] = f["filename"].rsplit(".", 1)[0] in converted_stems

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
MAC_FILL = PatternFill("solid", fgColor="E2EFDA")
NONMAC_FILL = PatternFill("solid", fgColor="FCE4D6")
THIN = Border(*[Side(style="thin", color="D9D9D9")] * 4)

def style_sheet(ws, widths, n_rows):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for c in ws[1]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2, max_row=n_rows + 1):
        for c in row:
            c.border = THIN
    ws.freeze_panes = "A2"

def first_author(author: str) -> str:
    if not author:
        return ""
    parts = [p.strip() for p in author.split(";") if p.strip()]
    if not parts:
        return ""
    fam = parts[0].split(",")[0].strip()
    return fam + " et al." if len(parts) > 1 else fam

wb = Workbook()

# ============ Sheet 1: 待转换文献清单 ============
ws1 = wb.active
ws1.title = "待转换文献清单"
ws1.append(["序号", "Zotero文件夹ID", "文件名", "CSV条目Key", "第一作者", "标题", "年份", "期刊", "DOI",
            "MAC相关", "判断依据", "大小MB", "转换方式"])
pending = [f for f in files if not f["conv_now"]]
for i, f in enumerate(pending, 1):
    size_mb = f["size_mb"]
    mode = "extract" if size_mb > 10 else "flash"
    ws1.append([i, f["folder"], f["filename"], f["csv_key"] or "", first_author(f.get("author") or ""),
                f["title"] or f["filename"], f["year"] or "", f["journal"] or "", f["doi"] or "",
                "是" if f["mac"] else "否", f["evidence"], size_mb, mode])
    r = ws1.max_row
    ws1.cell(r, 10).fill = MAC_FILL if f["mac"] else NONMAC_FILL
style_sheet(ws1, [6, 14, 60, 12, 22, 60, 8, 24, 26, 9, 22, 9, 10], len(pending))

# ============ Sheet 2: 全库分类汇报 ============
ws2 = wb.create_sheet("全库分类汇报")
ws2.append(["序号", "Zotero文件夹ID", "文件名", "CSV条目Key", "第一作者", "标题", "年份", "期刊", "DOI",
            "MAC相关", "判断依据", "是否在用户参考文献", "大小MB", "是否已转换", "是否入mucinous"])
for i, f in enumerate(files, 1):
    in_ref = "是" if f["in_user_refs"] else "否"
    ws2.append([i, f["folder"], f["filename"], f["csv_key"] or "", first_author(f.get("author") or ""),
                f["title"] or f["filename"], f["year"] or "", f["journal"] or "", f["doi"] or "",
                "是" if f["mac"] else "否", f["evidence"], in_ref, f["size_mb"],
                "是" if f["conv_now"] else "否", "是" if f["mac"] else "否"])
    r = ws2.max_row
    ws2.cell(r, 10).fill = MAC_FILL if f["mac"] else NONMAC_FILL
style_sheet(ws2, [6, 14, 60, 12, 22, 60, 8, 24, 26, 9, 22, 20, 9, 10, 12], len(files))

# ============ Sheet 3: 参考文献提及但库中无PDF ============
ws3 = wb.create_sheet("参考文献缺失文献")
ws3.append(["序号", "CSV条目Key", "标题", "年份", "DOI", "参考文献条目", "出处文档"])
missing = data["missing_refs"]
for i, m in enumerate(missing, 1):
    docs = "、".join(m.get("docs", []))
    ws3.append([i, m["key"], m["title"], m["year"], m["doi"] or "", m["ref"], docs])
# refs without CSV entry either
ref_dois = {(e["doi"] or "").lower() for e in data["refs"]}
found_dois = {(f["doi"] or "").lower() for f in files if f["doi"]}
no_pdf_refs = [e for e in data["refs"] if (e["doi"] or "").lower() not in found_dois and (e["doi"] or "").lower() not in {(m["doi"] or "").lower() for m in missing}]
for i, e in enumerate(no_pdf_refs, len(missing) + 1):
    ws3.append([i, "(索引中无此条目)", "", "", e["doi"] or "", e["raw"], "、".join(e["docs"])])
style_sheet(ws3, [6, 18, 70, 8, 26, 90, 26], len(missing) + len(no_pdf_refs))

# ============ Sheet 4: 处理说明 ============
ws4 = wb.create_sheet("处理说明")
notes = [
    ["分类规则", "1) 在 E:\\MAC\\202511前期资料整理 6 份文档参考文献清单中的文献 → MAC（必须入 mucinous）；2) 其余按标题判断，含 mucin/mucus/黏液 → MAC；3) 其余为非 MAC，不入 mucinous"],
    ["分类结果", f"共 {len(files)} 篇：MAC {sum(1 for f in files if f['mac'])} 篇，非MAC {sum(1 for f in files if not f['mac'])} 篇"],
    ["转换状态", f"已转换 {sum(1 for f in files if f['conv_now'])} 篇；待转换 {len(pending)} 篇（其中 >10MB 用 extract 模式 4 篇）"],
    ["ESGAR 2026", "PART I/II 为直肠癌影像通用共识，与黏液腺癌无直接关系 → 不入 mucinous（保留在 converted/）；converted/ 已有 ESGAR202601 PARTI/PARTII.md，本次跳过重复转换"],
    ["清理操作", "1) 删除空目录 papers/tmp-CE8PCs、papers/tmp-H5zCBz；2) 删除重复文献 papers/PXS25YBQ（与 4Q4STB45 MD5 完全一致）；3) 删除重复文件 papers/K9N4H2ML/ESGAR2026共识总结.md（与 CD84ZFXB 完全一致）"],
    ["疑似重复", "papers/XRLR7M9W/Melis 2010（小写 gene expression profiling）与 CL55BBYQ 同 DOI 但 MD5 不同，未转换，待用户确认"],
    ["参考文献缺失", f"{len(missing)} 篇在用户文档参考文献中但库中无 PDF，见『参考文献缺失文献』表；可据 DOI 在 Zotero 补充下载"],
    ["mucinous/", "MAC 相关文献的 converted markdown 副本存放于 mucinous/；非 MAC 文献不进入"],
]
ws4.append(["项目", "说明"])
for n in notes:
    ws4.append(n)
style_sheet(ws4, [16, 130], len(notes))

out = ROOT / "MAC文献分类与转换报告.xlsx"
wb.save(str(out))
print(f"[EXCEL] saved: {out}")
print(f"  pending={len(pending)} mac={sum(1 for f in files if f['mac'])} nonmac={sum(1 for f in files if not f['mac'])}")
print(f"  missing_refs={len(missing)} no_pdf_refs={len(no_pdf_refs)}")
